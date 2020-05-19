from openpyxl import load_workbook


class DataInfo:

    def __init__(self, prefix=''):
        self.file_set = []
        self.table_info = {}
        self.data_info = {}
        self.filedir = '服创大赛训练集-Inspur'
        self.filename = '训练集表格及字段说明.xlsx'
        self.prefix = prefix
        self.load()

    def load(self):
        wb = load_workbook(self.prefix + self.filedir + '/' + self.filename)
        ws = wb['脱敏后数据集']

        filenames = ws['B']
        filenames = [x.value.replace(' ', '') for i, x in enumerate(filenames)
                     if x.value is not None and x.value != '' and i != 0]
        self.file_set = [x.lower() for x in filenames]
        self.file_set.remove('enterprise_guarantee')  # 这个表格的数据处理还没想好 todo

        table_slice = 'B2:C' + str(ws.max_row)
        tables = ws[table_slice]
        tables = [(x[0].value, x[1].value) for x in tables if x[0].value is not None]
        for key, value in tables:
            self.table_info[key] = value

        data_slice = 'D2:E' + str(ws.max_row)
        datas = ws[data_slice]
        datas = [(x[0].value, x[1].value) for x in datas if x[0].value is not None]
        datas = [(x[0].replace('\n', '').replace(' ', ''),
                  x[1].replace('\n', '').replace(' ', ''))
                 for x in datas]
        for key, value in datas:
            self.data_info[key.lower()] = value


if __name__ == '__main__':
    datainfo = DataInfo()
    datainfo.load()
